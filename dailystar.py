#Daily Star News Site Scraping Program
#import all the required libraries
#check if file exist if not create a blank work book
#we want scraping to be store in an excel file
#get the current date
#crete a new sheet
#use iteration to store data
#save the workbook
#Script Written By : Ayan Upadhaya, contact: ayanU881@gmail.com

import os

import requests

import openpyxl

from bs4 import BeautifulSoup as bs

from datetime import datetime

"""WEB SCRAPING"""
base_link="https://www.thedailystar.net/english"
try:
	response=requests.get(base_link,timeout=10).content
except Exception as e:
	print(e)


soup=bs(response,"html.parser")
data=soup.find_all('h3',class_="title")


 
"""EXCEL AUTOMATION"""
file_name='star.xlsx'

all_files=os.listdir()

if file_name in all_files:
	wb=openpyxl.load_workbook(file_name)
else:
	wb=openpyxl.Workbook()

current=datetime.date(datetime.now())

wb.create_sheet(index=1,title=str(current))

sheet=wb[str(current)]

sheet['A1']="Title"
sheet['B1']="URL"

columns=['A','B']
#setting up column dimension
for cols in columns:
	sheet.column_dimensions[cols].width=35



i=2
for dt in data:
	title=dt.a.text
	url=dt.a.get('href')
	sheet['A'+str(i)]=title
	sheet['B'+str(i)]=base_link+url
	i+=1

wb.save(file_name)

print("Success!")
	
